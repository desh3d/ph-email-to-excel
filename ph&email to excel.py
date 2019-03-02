import openpyxl, os, re, pyperclip

# This program xtracts all text that matchs an email id or phone number pattern from clipboard  and writes them directly to an excel file!!!!
# Its useful for the right person



os.chdir('C:\\Users\\dhira\\Desktop\\mywork')

wb = openpyxl.Workbook()
sheet = wb['Sheet']
wb.create_sheet(index=2, title='Sheet2')

text = pyperclip.paste()    # text data to search from   

phoneRegex = re.compile(r'''(
    (\+?\d\d|0)?                 #+91
    (\d{3}|\(\d{3}\))            # area code
    (\s|-|\.)?                   # separator
    \d{3}                        # first 3 digits
    (\s|-|\.)?                   # separator
    \d{4}                        # last 4 digits
    (\s*(ext|x|ext.)\s*\d{2,5})? # extension
    )''', re.VERBOSE)

emailregex = re.compile(r'''(
    [a-zA-Z0-9._%+-]+       #username
    \@                      #@
    [a-zA-Z0-9.-]+          #domain
    (\.[a-zA-Z]{2,4})       #.com
    )''', re.VERBOSE)

mo = phoneRegex.findall(text)                      
mo2 = emailregex.findall(text)

numberslist= []
emailist = []

for i in range(len(mo)):
    numbers= (mo[i][0])
    #print(numbers)
    numberslist.append(numbers)
for j in range(len(mo2)):
    emails= (mo2[j][0])
    #print(emails)
    emailist.append(emails)

lister = numberslist + emailist

''' SAMPLE  my phone no is +918975256588 and also 09890769223 while my dad's phone number is 8087760723 and american no is 123-152.1542 or -754-8245,which can also be written
    as (123)-453-7895 or +218975256588 or 08975256588my email is dhirajdeshpande1596@gmail.com while jimmisys email id is divase.akshay@gmail.comdhirajdeshpande1596@gmail.com
       divase.akshay@gmail.com  jimmy.fallon96@gmail.com  conanobrian@reddifmail.com jimmy.fallon96@gmail.com. The conanobrian@reddifmail.com is conan\'s email id'''

print('copied to clipboard')
#print(numberslist)
#print(emailist)
#print(lister)
for nums in range(len(numberslist)):
    sheet.cell(row=nums+1, column=1, value=numberslist[nums])
    sheet.cell(row=nums+1, column=2, value=emailist[nums])

wb.save('example3.xlsx')
#pyperclip.copy('  '.join(lister))
#print(mo)
#print(mo2)
